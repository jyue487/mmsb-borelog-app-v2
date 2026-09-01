"""Reads a generated workbook the way the borelog report's Python does, and prints it.

    python3 packages/ags-excel/scripts/verify.py [workbook.xlsx]

The report loads with `openpyxl.load_workbook(data_only=True)`, which never evaluates a
formula -- it returns the value Excel last cached. So this is the check that matters: it
sees exactly what the report will see, including whether the caches this exporter writes
for the SPT sheet's N value and reported result actually took.
"""
import sys
import openpyxl

DEFAULT = 'packages/ags-excel/out/fixture.xlsx'
path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT

wb = openpyxl.load_workbook(filename=path, data_only=True)


def rows(sheet_name, columns, start_row):
    """Every row until column B runs out -- the same stop condition the parsers use."""
    ws = wb[sheet_name]
    out = []
    row = start_row
    while ws[f'B{row}'].value is not None:
        out.append([ws[f'{c}{row}'].value for c in columns])
        row += 1
    return out


def show(title, columns, values):
    print(f'\n### {title}  ({len(values)} rows)')
    print('    ' + ' | '.join(c.ljust(10) for c in columns))
    for value in values:
        print('    ' + ' | '.join(('' if v is None else str(v))[:10].ljust(10) for v in value))


ws = wb['Project - AGS']
print('### Project')
for ref, label in [('I4', 'PROJ_ID'), ('D4', 'name'), ('D6', 'client'),
                   ('D7', 'engineer'), ('D8', 'location'), ('D9', 'contractor')]:
    print(f'    {label:<10} {ws[ref].value!r}')

show('Holes', ['PROJ_ID', 'HOLE_ID', 'TYPE', 'NATE', 'NATN', 'FDEP', 'GL', 'STAR', 'ENDD', 'BACD', 'LOG'],
     rows('Holes - AGS', 'ABCDEFGHIJK', 6))
show('Progress', ['PROJ_ID', 'HOLE_ID', 'DATE', 'TIME', 'DEP', 'CAS', 'WAT'],
     rows('Progress - AGS', 'ABCDEFG', 6))
show('SPT', ['PROJ_ID', 'HOLE_ID', 'TOP', 'NVAL', 'REP'],
     [[r[0], r[1], r[2], r[3], r[4]] for r in rows('SPT - AGS', 'ABCST', 7)])
show('Geology', ['PROJ_ID', 'HOLE_ID', 'GEOL', 'LEG', 'TOP', 'BASE', 'DESC'],
     rows('Geology - AGS', 'ABCDEFG', 6))
show('Samples', ['PROJ_ID', 'HOLE_ID', 'TOP', 'BASE', 'TYPE', 'REF', 'DIA', 'REM'],
     rows('Samples - AGS', 'ABCDEFGH', 6))
show('Core', ['PROJ_ID', 'HOLE_ID', 'TOP', 'BASE', 'PREC', 'SREC', 'RQD', 'DIAM'],
     rows('Core - AGS', 'ABCDEFGH', 6))
show('Water Strike', ['PROJ_ID', 'HOLE_ID', 'DEP', 'DATE', 'TIME', 'CAS'],
     rows('Water Strike - AGS', 'ABCDEF', 6))
